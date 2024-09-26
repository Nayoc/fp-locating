from openpyxl import load_workbook
from openpyxl import Workbook


def read_excel(path):
    wb = load_workbook(filename=path)
    ws = wb.active
    data = []
    for row in ws.iter_rows(values_only=True):
        r = list(row)
        data.append(r)
    return data


# 写入 Excel 文件，写入整个List
def write_excel(data, output_file):
    wb = Workbook()
    ws = wb.active
    for row in data:
        ws.append(row)
    wb.save_model(output_file)


# 写入 Excel 文件,写入指定行
def write_excel(data, row, output_file):
    wb = load_workbook(filename=output_file)
    sheet = wb['Sheet1']

    for col_index, value in enumerate(data, start=1):
        sheet.cell(row=row + 1, column=col_index, value=value)

    wb.save_model(output_file)
